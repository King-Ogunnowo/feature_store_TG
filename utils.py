from openpyxl import load_workbook
from collections import Counter
import pandas as pd
import requests
import datetime
import json

"""
This module contains functions used to create feature store
"""

MAX_LEN = 18
CATEGORIES = [
    'Charges', 'POS', 'WEB', 'ATM', 
    'Airtime_Data', 'Transfer', 'MISC'
]

def get_sheetnames_xlsx(filepath):
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    return wb.sheetnames

def keywords_extractor(keywords_df):
    """
    Takes a keywords dataframe and extracts 
    its related keywords as a dictionary of 
    keywords
    """
    key_word_corpus = []
    for row in list(keywords_df.to_dict().values()):
        if all([isinstance(val, float) for val in row.values()]):
            pass
        else:
            key_words = filter(
                lambda key: not isinstance(key, float), 
                list(row.values())[:MAX_LEN]
            )
            lower_cased = list(map(lambda key: key.lower(), key_words))
            key_word_corpus.append(lower_cased)
    
    return {
        cat: keyword_list[1:]
        for cat, keyword_list in zip(CATEGORIES, key_word_corpus)
    }

def hash_list(list_type):
    """
    Combines several list into one
    """
    
    result = []
    for _list in list_type:
        result += _list
    
    return result

def merge_sheet_observations(*sheets):
    """
    Merges the observed keywords from 
    the various sheets and merges them 
    into one.
    """
    
    merged_results = {}
    for cat in CATEGORIES:
        merged_results[cat] = list(set(
            hash_list([
                sheet[cat] 
                for sheet in sheets
            ])
        ))
    
    return merged_results

def get_categories(text, dictionary):
    """
    Function to get categories based on keywords.
    """
    categories = []
    for i in range(len(dictionary)):
        for item in text:
            if item in list(dictionary.values())[i]:
                categories.append((list(dictionary.keys())[i]))           
    return categories

def send_status_msg(slack_msg):
    """
    function to notify slack channel about insight extraction
    """
    whurl = 'https://hooks.slack.com/services/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'
    requests.post(whurl,data=json.dumps(slack_msg))
    
def Convert(month):
    if type(month) != int:
        if '[' in month:
            return re.findall('[0-9]', month)
    else:
        return month
    
def get_previous_month():
    """function to get month and year values of previous month"""
    first_date_of_curr_month = datetime.datetime.today().replace(day = 1)
    one_day_back = first_date_of_curr_month - datetime.timedelta(days = 1)
    return one_day_back.month, one_day_back.year
    


